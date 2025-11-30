#!/usr/bin/env python3
"""
Módulo de Exportación Excel para Planillas de Producción
Optimizador de Producción v1.3.3
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import os
from typing import List, Dict, Optional

class ExportadorExcel:
    """Clase para generar archivos Excel de planillas de producción"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def generar_planilla_produccion(self, programacion: Dict, tareas: List[Dict], output_path: str) -> bool:
        """
        Generar Excel de planilla de producción semanal
        
        Args:
            programacion: Datos de la programación
            tareas: Lista de tareas programadas
            output_path: Ruta donde guardar el Excel
            
        Returns:
            bool: True si se generó correctamente
        """
        try:
            # Crear workbook
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Planilla de Producción"
            
            # Configurar estilos
            self._configurar_estilos()
            
            # Agregar información de la programación
            self._agregar_cabecera_programacion(programacion)
            
            # Agregar tabla de tareas
            self._agregar_tabla_tareas(tareas)
            
            # Agregar resumen por máquina
            self._agregar_resumen_maquinas(tareas)
            
            # Agregar métricas
            self._agregar_metricas(programacion, tareas)
            
            # Guardar archivo
            self.workbook.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error generando Excel de planilla: {e}")
            return False
    
    def _configurar_estilos(self):
        """Configurar estilos para el Excel"""
        # Estilos de fuente
        self.font_titulo = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        self.font_subtitulo = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.font_cabecera = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        self.font_normal = Font(name='Arial', size=10)
        
        # Estilos de relleno
        self.fill_titulo = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.fill_subtitulo = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        self.fill_cabecera = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.fill_alterno = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Estilos de alineación
        self.align_centro = Alignment(horizontal='center', vertical='center')
        self.align_izquierda = Alignment(horizontal='left', vertical='center')
        
        # Estilos de borde
        self.borde_fino = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _agregar_cabecera_programacion(self, programacion: Dict):
        """Agregar cabecera con información de la programación"""
        # Título principal
        self.worksheet['A1'] = "PLANILLA DE PRODUCCIÓN SEMANAL"
        self.worksheet['A1'].font = self.font_titulo
        self.worksheet['A1'].fill = self.fill_titulo
        self.worksheet['A1'].alignment = self.align_centro
        self.worksheet.merge_cells('A1:H1')
        
        # Información de la programación
        fila = 3
        info_programacion = [
            ["Semana de Producción:", f"Semana {programacion['semana_produccion']} - {programacion['anio']}"],
            ["Estado:", programacion['estado'].value if hasattr(programacion['estado'], 'value') else str(programacion['estado'])],
            ["Fecha de Creación:", programacion['fecha_creacion'].strftime("%d/%m/%Y %H:%M") if programacion['fecha_creacion'] else "N/A"],
            ["Objetivo:", programacion['objetivo_usado']],
            ["Makespan Planificado:", f"{programacion['makespan_planificado']:.1f} minutos" if programacion['makespan_planificado'] else "N/A"]
        ]
        
        if programacion.get('aprobada_por'):
            info_programacion.append(["Aprobada por:", programacion['aprobada_por']])
        if programacion.get('fecha_aprobacion'):
            info_programacion.append(["Fecha de Aprobación:", programacion['fecha_aprobacion'].strftime("%d/%m/%Y %H:%M")])
        
        for etiqueta, valor in info_programacion:
            self.worksheet[f'A{fila}'] = etiqueta
            self.worksheet[f'A{fila}'].font = self.font_cabecera
            self.worksheet[f'A{fila}'].fill = self.fill_cabecera
            self.worksheet[f'A{fila}'].alignment = self.align_izquierda
            
            self.worksheet[f'B{fila}'] = valor
            self.worksheet[f'B{fila}'].font = self.font_normal
            self.worksheet[f'B{fila}'].alignment = self.align_izquierda
            
            fila += 1
        
        # Espacio
        fila += 1
        return fila
    
    def _agregar_tabla_tareas(self, tareas: List[Dict]):
        """Agregar tabla de tareas programadas"""
        # Encontrar la fila donde empezar
        fila_inicio = 10
        
        # Cabeceras de la tabla
        cabeceras = ["#", "Trabajo", "Tarea", "Máquina", "Operador", "Inicio", "Fin", "Duración (min)"]
        
        for col, cabecera in enumerate(cabeceras, 1):
            celda = self.worksheet.cell(row=fila_inicio, column=col)
            celda.value = cabecera
            celda.font = self.font_cabecera
            celda.fill = self.fill_cabecera
            celda.alignment = self.align_centro
            celda.border = self.borde_fino
        
        # Datos de las tareas
        for i, tarea in enumerate(tareas, 1):
            fila = fila_inicio + i
            
            datos = [
                i,
                tarea.get('trabajo_nombre', tarea.get('trabajo_id', 'N/A')),
                tarea.get('nombre', 'N/A'),
                f"M{tarea['maquina_id']}",
                tarea.get('operador_id', 'Por asignar'),
                self._formatear_hora(tarea['inicio_planificado']),
                self._formatear_hora(tarea['fin_planificado']),
                tarea['duracion_planificada']
            ]
            
            for col, dato in enumerate(datos, 1):
                celda = self.worksheet.cell(row=fila, column=col)
                celda.value = dato
                celda.font = self.font_normal
                celda.alignment = self.align_centro if col > 2 else self.align_izquierda
                celda.border = self.borde_fino
                
                # Relleno alternado
                if i % 2 == 0:
                    celda.fill = self.fill_alterno
        
        # Ajustar ancho de columnas
        self._ajustar_ancho_columnas()
        
        return fila_inicio + len(tareas) + 2
    
    def _agregar_resumen_maquinas(self, tareas: List[Dict]):
        """Agregar resumen por máquina"""
        # Encontrar la fila donde empezar
        fila_inicio = 15 + len(tareas)
        
        # Título del resumen
        self.worksheet[f'A{fila_inicio}'] = "RESUMEN POR MÁQUINA"
        self.worksheet[f'A{fila_inicio}'].font = self.font_subtitulo
        self.worksheet[f'A{fila_inicio}'].fill = self.fill_subtitulo
        self.worksheet[f'A{fila_inicio}'].alignment = self.align_centro
        self.worksheet.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
        
        # Cabeceras del resumen
        fila_inicio += 2
        cabeceras = ["Máquina", "Tareas", "Tiempo Total (min)", "Tiempo Total (h)"]
        
        for col, cabecera in enumerate(cabeceras, 1):
            celda = self.worksheet.cell(row=fila_inicio, column=col)
            celda.value = cabecera
            celda.font = self.font_cabecera
            celda.fill = self.fill_cabecera
            celda.alignment = self.align_centro
            celda.border = self.borde_fino
        
        # Agrupar tareas por máquina
        tareas_por_maquina = {}
        for tarea in tareas:
            maquina = tarea['maquina_id']
            if maquina not in tareas_por_maquina:
                tareas_por_maquina[maquina] = []
            tareas_por_maquina[maquina].append(tarea)
        
        # Datos del resumen
        fila = fila_inicio + 1
        for maquina in sorted(tareas_por_maquina.keys()):
            tareas_maq = tareas_por_maquina[maquina]
            tiempo_total = sum(t['duracion_planificada'] for t in tareas_maq)
            tiempo_horas = tiempo_total / 60
            
            datos = [
                f"M{maquina}",
                len(tareas_maq),
                tiempo_total,
                f"{tiempo_horas:.1f}"
            ]
            
            for col, dato in enumerate(datos, 1):
                celda = self.worksheet.cell(row=fila, column=col)
                celda.value = dato
                celda.font = self.font_normal
                celda.alignment = self.align_centro
                celda.border = self.borde_fino
                
                # Relleno alternado
                if fila % 2 == 0:
                    celda.fill = self.fill_alterno
            
            fila += 1
        
        return fila + 2
    
    def _agregar_metricas(self, programacion: Dict, tareas: List[Dict]):
        """Agregar métricas y estadísticas"""
        # Encontrar la fila donde empezar
        fila_inicio = 20 + len(tareas)
        
        # Título de métricas
        self.worksheet[f'A{fila_inicio}'] = "MÉTRICAS Y ESTADÍSTICAS"
        self.worksheet[f'A{fila_inicio}'].font = self.font_subtitulo
        self.worksheet[f'A{fila_inicio}'].fill = self.fill_subtitulo
        self.worksheet[f'A{fila_inicio}'].alignment = self.align_centro
        self.worksheet.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
        
        # Calcular métricas
        total_tareas = len(tareas)
        total_tiempo = sum(t['duracion_planificada'] for t in tareas)
        tiempo_promedio = total_tiempo / total_tareas if total_tareas > 0 else 0
        maquinas_utilizadas = len(set(t['maquina_planificada'] for t in tareas))
        
        # Métricas
        fila_inicio += 2
        metricas = [
            ["Total de Tareas:", total_tareas],
            ["Total de Tiempo (min):", total_tiempo],
            ["Total de Tiempo (h):", f"{total_tiempo/60:.1f}"],
            ["Tiempo Promedio por Tarea (min):", f"{tiempo_promedio:.1f}"],
            ["Máquinas Utilizadas:", maquinas_utilizadas],
            ["Makespan Planificado (min):", f"{programacion['makespan_planificado']:.1f}" if programacion['makespan_planificado'] else "N/A"]
        ]
        
        for i, (etiqueta, valor) in enumerate(metricas):
            fila = fila_inicio + i
            
            self.worksheet[f'A{fila}'] = etiqueta
            self.worksheet[f'A{fila}'].font = self.font_cabecera
            self.worksheet[f'A{fila}'].fill = self.fill_cabecera
            self.worksheet[f'A{fila}'].alignment = self.align_izquierda
            
            self.worksheet[f'B{fila}'] = valor
            self.worksheet[f'B{fila}'].font = self.font_normal
            self.worksheet[f'B{fila}'].alignment = self.align_izquierda
    
    def _ajustar_ancho_columnas(self):
        """Ajustar el ancho de las columnas"""
        anchos = {
            'A': 5,   # #
            'B': 25,  # Trabajo
            'C': 20,  # Tarea
            'D': 10,  # Máquina
            'E': 15,  # Operador
            'F': 10,  # Inicio
            'G': 10,  # Fin
            'H': 12   # Duración
        }
        
        for col, ancho in anchos.items():
            self.worksheet.column_dimensions[col].width = ancho
    
    def _formatear_hora(self, minutos_desde_inicio: Optional[int]) -> str:
        """Convertir minutos desde inicio a formato HH:MM"""
        if minutos_desde_inicio is None:
            return "N/A"
        
        horas = minutos_desde_inicio // 60
        minutos = minutos_desde_inicio % 60
        return f"{horas:02d}:{minutos:02d}"
    
    def generar_csv_sistemas_externos(self, programacion: Dict, tareas: List[Dict], output_path: str) -> bool:
        """
        Generar CSV para sistemas externos (ERP/MES)
        
        Args:
            programacion: Datos de la programación
            tareas: Lista de tareas programadas
            output_path: Ruta donde guardar el CSV
            
        Returns:
            bool: True si se generó correctamente
        """
        try:
            # Preparar datos para CSV
            datos_csv = []
            
            for tarea in tareas:
                fila = {
                    'semana_produccion': programacion['semana_produccion'],
                    'anio': programacion['anio'],
                    'programacion_id': programacion['id'],
                    'tarea_id': tarea.get('id', 'N/A'),
                    'trabajo_nombre': tarea.get('trabajo_nombre', tarea.get('trabajo_id', 'N/A')),
                    'tarea_nombre': tarea.get('nombre', 'N/A'),
                    'maquina': tarea['maquina_id'],
                    'operador': tarea.get('operador_id', ''),
                    'inicio_planificado': self._formatear_hora(tarea['inicio_planificado']),
                    'fin_planificado': self._formatear_hora(tarea['fin_planificado']),
                    'duracion_planificada': tarea['duracion_planificada'],
                    'prioridad': tarea.get('prioridad', 'Normal'),
                    'estado_programacion': programacion['estado'].value if hasattr(programacion['estado'], 'value') else str(programacion['estado']),
                    'fecha_creacion': programacion['fecha_creacion'].strftime("%Y-%m-%d %H:%M:%S") if programacion['fecha_creacion'] else '',
                    'objetivo_usado': programacion['objetivo_usado'],
                    'makespan_planificado': programacion['makespan_planificado'] if programacion['makespan_planificado'] else ''
                }
                datos_csv.append(fila)
            
            # Crear DataFrame y guardar CSV
            df = pd.DataFrame(datos_csv)
            df.to_csv(output_path, index=False, encoding='utf-8')
            
            return True
            
        except Exception as e:
            print(f"Error generando CSV: {e}")
            return False
